import base64
import csv
import io
from datetime import date, datetime

from openpyxl import load_workbook

from odoo import _, api, fields, models
from odoo.exceptions import UserError

# Template headers (English source) and their zh_CN counterparts from
# i18n/zh_CN.po — both are accepted when parsing uploaded files.
IMPORT_HEADERS = [
    ('equipment', 'Equipment Code', '设备编码'),
    ('date', 'Date', '日期'),
    ('shift', 'Shift', '班次'),
    ('planned_time', 'Planned Working Time (h)', '计划工作时间(小时)'),
    ('downtime_hours', 'Downtime (h)', '停机时间(小时)'),
    ('design_capacity', 'Design Capacity (pcs/h)', '设计产能(件/小时)'),
    ('actual_output', 'Actual Output (pcs)', '实际产量(件)'),
    ('qualified_qty', 'Qualified Qty (pcs)', '合格品数量(件)'),
]

HEADER_ALIASES = {
    field: {en.strip().lower(), zh.strip()} for field, en, zh in IMPORT_HEADERS
}

SHIFT_ALIASES = {
    'all': 'all', 'day': 'day', 'night': 'night',
    'all day': 'all', 'day shift': 'day', 'night shift': 'night',
    '全天': 'all', '白班': 'day', '夜班': 'night',
}

FLOAT_FIELDS = (
    'planned_time', 'downtime_hours', 'design_capacity',
    'actual_output', 'qualified_qty',
)


class OeeImportWizard(models.TransientModel):
    """Import OEE records from the xlsx/csv template.

    Rows whose equipment, date and shift match an existing record replace
    that record (or are rejected when overwrite is unchecked).
    """
    _name = 'sn.wsd.device.oee.import.wizard'
    _description = 'OEE Import'

    file = fields.Binary(string='File', required=True, filename='file_name')
    file_name = fields.Char(string='File Name')
    overwrite = fields.Boolean(
        string='Overwrite Existing Records', default=True,
        help='Rows matching an existing record (same equipment, date and '
             'shift) replace it instead of being rejected.')

    def action_import_records(self):
        self.ensure_one()
        rows = self._parse_file()
        record_model = self.env['sn.wsd.device.oee.record']
        errors = []
        equipment_cache = {}
        vals_list = []
        for row_number, row in rows:
            error = self._validate_row(
                row, row_number, equipment_cache)
            if error:
                errors.append(error)
                continue
            vals_list.append((row_number, row))
        if errors:
            raise UserError('\n'.join(errors))

        created = updated = 0
        for row_number, row in vals_list:
            existing = record_model.search([
                ('equipment_id', '=', row['equipment_id'].id),
                ('date', '=', row['date']),
                ('shift', '=', row['shift']),
            ], limit=1)
            if existing:
                if not self.overwrite:
                    errors.append(_(
                        'Row %(row)s: a record already exists for this '
                        'equipment, date and shift.',
                        row=row_number))
                    continue
                existing.write(self._record_vals(row))
                updated += 1
            else:
                record_model.create(self._record_vals(row))
                created += 1
        if errors:
            raise UserError('\n'.join(errors))

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'type': 'success',
                'title': _('OEE Records Imported'),
                'message': _(
                    '%(created)s record(s) created, %(updated)s record(s) '
                    'updated.', created=created, updated=updated),
                'next': {'type': 'ir.actions.act_window_close'},
            },
        }

    def _record_vals(self, row):
        return {
            'equipment_id': row['equipment_id'].id,
            'company_id': row['equipment_id'].company_id.id,
            'date': row['date'],
            'shift': row['shift'],
            'planned_time': row['planned_time'],
            'downtime_hours': row['downtime_hours'],
            'design_capacity': row['design_capacity'],
            'actual_output': row['actual_output'],
            'qualified_qty': row['qualified_qty'],
            'state': 'done',
        }

    def _validate_row(self, row, row_number, equipment_cache):
        """Return an error message (translated, row-prefixed) or None."""
        key = row['equipment']
        if key not in equipment_cache:
            equipment = self.env['sn.wsd.device.equipment'].search([
                '|', ('code', '=ilike', key), ('name', '=ilike', key),
            ], limit=2)
            if not equipment:
                equipment_cache[key] = _(
                    'Row %(row)s: equipment "%(value)s" not found.',
                    row=row_number, value=key)
            elif len(equipment) > 1:
                equipment_cache[key] = _(
                    'Row %(row)s: equipment "%(value)s" matches several '
                    'records, use its code.',
                    row=row_number, value=key)
            else:
                equipment_cache[key] = equipment[0]
        cached = equipment_cache[key]
        if isinstance(cached, str):
            return cached
        row['equipment_id'] = cached

        planned_time = row['planned_time']
        downtime_hours = row['downtime_hours']
        design_capacity = row['design_capacity']
        actual_output = row['actual_output']
        qualified_qty = row['qualified_qty']
        if planned_time <= 0:
            return _('Row %(row)s: planned working time must be positive.',
                     row=row_number)
        if downtime_hours < 0 or downtime_hours > planned_time:
            return _(
                'Row %(row)s: downtime must be between 0 and the planned '
                'working time.', row=row_number)
        if design_capacity <= 0:
            return _('Row %(row)s: design capacity must be positive.',
                     row=row_number)
        if actual_output < 0:
            return _('Row %(row)s: actual output cannot be negative.',
                     row=row_number)
        if qualified_qty < 0 or qualified_qty > actual_output:
            return _(
                'Row %(row)s: qualified qty must be between 0 and the '
                'actual output.', row=row_number)
        return None

    def _parse_file(self):
        """Return [(row_number, {field: value}), ...] from xlsx or csv."""
        self.ensure_one()
        name = (self.file_name or '').lower()
        data = base64.b64decode(self.file)
        if name.endswith('.csv'):
            text = data.decode('utf-8-sig')
            raw_rows = list(csv.reader(io.StringIO(text)))
        elif name.endswith('.xlsx'):
            try:
                workbook = load_workbook(
                    io.BytesIO(data), read_only=True, data_only=True)
            except Exception:
                raise UserError(_(
                    'Could not read the file. Please upload a valid xlsx '
                    'generated from the template.'))
            raw_rows = [
                list(row) for row in
                workbook.worksheets[0].iter_rows(values_only=True)]
        else:
            raise UserError(_(
                'Unsupported file format. Upload an xlsx or csv file.'))
        return self._parse_rows(raw_rows)

    def _parse_rows(self, raw_rows):
        # Map normalized header cells to import fields.
        columns = {}
        for row in raw_rows:
            cells = [str(v).strip() if v is not None else ''
                     for v in row]
            if not any(cells):
                continue
            for index, cell in enumerate(cells):
                for field, aliases in HEADER_ALIASES.items():
                    if cell.lower() in aliases:
                        columns[field] = index
            if 'equipment' in columns and 'date' in columns:
                break
        if 'equipment' not in columns or 'date' not in columns:
            raise UserError(_(
                'The file is missing its header row. Use the download '
                'template.'))

        rows = []
        header_seen = False
        for row_number, row in enumerate(raw_rows, start=1):
            cells = [str(v).strip() if v is not None else ''
                     for v in row]
            if not any(cells):
                continue
            is_header = any(
                cell.lower() in aliases
                for cell in cells
                for aliases in HEADER_ALIASES.values())
            if is_header and not header_seen:
                header_seen = True
                continue
            if is_header:
                continue

            def cell(field):
                index = columns.get(field)
                return cells[index] if index is not None and \
                    index < len(cells) else ''

            equipment = cell('equipment')
            if not equipment or equipment.startswith('Example:'):
                continue
            raw_date = cell('date')
            try:
                parsed_date = self._parse_date(raw_date)
            except (ValueError, TypeError):
                raise UserError(_(
                    'Row %(row)s: invalid date "%(value)s" (expected '
                    'YYYY-MM-DD).', row=row_number, value=raw_date))
            shift_raw = cell('shift') or 'all'
            shift = SHIFT_ALIASES.get(shift_raw.strip().lower())
            if shift is None:
                raise UserError(_(
                    'Row %(row)s: invalid shift "%(value)s".',
                    row=row_number, value=cell('shift')))
            values = {
                'equipment': equipment,
                'date': parsed_date,
                'shift': shift,
            }
            defaults = {
                'planned_time': 8.0, 'downtime_hours': 0.0,
                'design_capacity': 0.0, 'actual_output': 0,
                'qualified_qty': 0,
            }
            for field in FLOAT_FIELDS:
                raw_value = cell(field)
                try:
                    values[field] = float(raw_value) if raw_value \
                        else defaults[field]
                except ValueError:
                    raise UserError(_(
                        'Row %(row)s: invalid number "%(value)s".',
                        row=row_number, value=raw_value))
            rows.append((row_number, values))
        if not rows:
            raise UserError(_('No data row found in the file.'))
        return rows

    @api.model
    def _parse_date(self, value):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return fields.Date.to_date(str(value).strip())
