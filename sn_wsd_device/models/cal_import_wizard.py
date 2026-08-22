import io

from openpyxl import load_workbook

from odoo import _, api, fields, models
from odoo.exceptions import UserError
from odoo.fields import Command


class CalibrationLineImportWizard(models.TransientModel):
    """Upload the Excel template, preview the parsed rows, then write
    them into the task lines on confirmation."""
    _name = 'sn.wsd.device.cal.line.import.wizard'
    _description = 'Calibration Line Import'

    task_id = fields.Many2one(
        'sn.wsd.device.cal.task', string='Calibration Task', required=True)
    file = fields.Binary(string='File', required=True)
    file_name = fields.Char(string='File Name')
    preview_line_ids = fields.One2many(
        'sn.wsd.device.cal.line.import.wizard.line', 'wizard_id',
        string='Preview Lines')

    def action_parse(self):
        self.ensure_one()
        self.preview_line_ids.unlink()
        rows = self._parse_xlsx()
        self.write({'preview_line_ids': [
            Command.create({'item_name': item, 'before_value': before,
                            'after_value': after, 'line_note': note})
            for item, before, after, note in rows]})
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'sn.wsd.device.cal.line.import.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_import(self):
        self.ensure_one()
        self.task_id.write({'line_ids': [
            Command.create({
                'item_name': line.item_name,
                'before_value': line.before_value,
                'after_value': line.after_value,
                'line_note': line.line_note,
            }) for line in self.preview_line_ids]})
        return {'type': 'ir.actions.act_window_close'}

    def _parse_xlsx(self):
        import base64
        try:
            workbook = load_workbook(
                io.BytesIO(base64.b64decode(self.file)),
                read_only=True, data_only=True)
        except Exception:
            raise UserError(_(
                'Could not read the file. Please upload a valid xlsx '
                'generated from the template.'))
        sheet = workbook.worksheets[0]
        rows = []
        header = None
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            values = ['' if v is None else str(v).strip() for v in row[:4]]
            if not any(values):
                continue
            if index == 0:
                header = values
                continue
            if header and header[0] == 'Example: voltage measurement':
                continue
            rows.append(tuple(values))
        # Skip the example row that follows the header.
        rows = [row for row in rows
                if row[0] != 'Example: voltage measurement']
        if not rows:
            raise UserError(_('No data row found in the file.'))
        return rows


class CalibrationLineImportWizardLine(models.TransientModel):
    _name = 'sn.wsd.device.cal.line.import.wizard.line'
    _description = 'Calibration Line Import Preview Row'

    wizard_id = fields.Many2one(
        'sn.wsd.device.cal.line.import.wizard', required=True,
        ondelete='cascade')
    item_name = fields.Char(string='Item Name')
    before_value = fields.Char(string='Before Calibration Value')
    after_value = fields.Char(string='After Calibration Value')
    line_note = fields.Char(string='Line Note')
